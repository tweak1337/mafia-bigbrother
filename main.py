import time

from aiogram import Dispatcher, Bot, executor, types

import fuck
import os
from datetime import datetime, timedelta
import openpyxl
import asyncio
import aioschedule
import random
import re
import json
import pymorphy2
from googletrans import Translator

import parser
from parser import *

bot = Bot(token=os.getenv('TOKEN'))

dp = Dispatcher(bot)

m = pymorphy2.MorphAnalyzer()

message_counter = 0


@dp.message_handler(content_types=['new_chat_members'])
async def user_join(message: types.Message):

    try:
        user = message.new_chat_members[0].username
    except:
        user = message.new_chat_members[0].first_name

    await message.answer(f'Привет @{user}!\nДобро пожаловать к нам в МафКлуб🤗🕵🏻‍♂️\n'
                         f'Здесь ты сможешь найти друзей, а так же быть в курсе всех предстоящих игр и мероприятий😉\n'
                         f'Но у нас есть простые правила:\n'
                         f'1. В чате запрещается выражаться матом (скрытый мат допускается)🤬\n'
                         f'2. Переходить на личности и оскорблять собеседника тоже нельзя, мы - живем дружно🫱🏻‍🫲🏼\n'
                         f'3. Сливать личные переписки, фото и видео содержащие оскорбительный характер оппонента🫣\n'
                         f'Очень рады, что ты к нам присоединился❤️')


@dp.message_handler(commands=['help', 'start'])
async def user_join(message: types.Message):
    await message.answer(f'Привет! Я бот этой группы. В основном я цензурирую мат, но еще я умею оскорблять, и присылать мемы\n\n'
                         f'Полный список команд доступен при вводе "/" в поле для ввода текста.')


@dp.message_handler(commands=['оскорбить', 'insult'])
async def user_join(message: types.Message):
    try:
        try:
            user = message.from_user.username
        except:
            user = message.from_user.first_name
    except:
        user = 'Unnamed'

    file = 'story.xlsx'
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    range_cells = sheet['B2':f'B{row_count}']
    insult_list = []
    for row in range_cells:
        for cell in row:
            insult_list.append(cell.value)

    insult = random.choice(insult_list)

    try:
        await message.answer(f'@{message.reply_to_message.from_user.username}, {insult}')
    except:
        await message.answer(insult)



@dp.message_handler(commands=['сохранить', 'save'])
async def user_join(message: types.Message):

    try:
        try:
            user = message.from_user.username
        except:
            user = message.from_user.first_name
    except:
        user = 'Unnamed'

    file = 'story.xlsx'
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row

    if '/save' in str(message.text):
        insult = message.text[6:]
    else:
        insult = message.text[11:]
    sheet[f'A{row_count + 1}'] = str(user)
    sheet[f'B{row_count + 1}'] = str(insult)
    wb.save(file)

    text = insult.split()
    text = [re.sub(r'[^\w\s]', '', i) for i in text]
    counter = 0
    for word in fuck.fuck_list:
        for mes in text:
            if word == mes:
                counter +=1

    if counter > 0:
        await message.delete()
        await message.answer(f'Маты в оскорблениях запрещены, как бы странно это не звучало')
    else:
        try:
            await message.answer(f'Оскорбление "{insult}" сохранено.')
        except:
            await message.answer(insult)


def normalize(s1):
    sent = []
    stroka = s1.replace('?', '')
    stroka = stroka.replace('!', '')
    stroka = stroka.replace('.', '')
    stroka = stroka.replace(',', '')

    s = stroka.lower().split()

    for i in s:
        p = m.parse(i)[0]
        sent.append(p.normal_form)
    s2 = ' '.join(sent)

    return s2

@dp.message_handler(commands=['translate', 'перевести'])
async def user_join(message: types.Message):

    if not message.reply_to_message:
        await message.answer("Нужно процитировать сообщение, чтобы его перевести.")
        return
    else:
        translator = Translator()
        translate = translator.translate(text=message.reply_to_message.text, dest='ru')
        translate = translate.text
        await message.answer(translate, reply=True)


@dp.message_handler(commands=['translate_eng'])
async def user_join(message: types.Message):

    if not message.reply_to_message:
        await message.answer("Нужно процитировать сообщение, чтобы его перевести.")
        return
    else:
        translator = Translator()
        translate = translator.translate(text=message.reply_to_message.text, dest='en')
        translate = translate.text
        await message.answer(translate, reply=True)

@dp.message_handler(commands=['meme'])
async def user_join(message: types.Message):

    parser.whole_memes()
    photo = open('my_image.jpg', 'rb')
    await bot.send_photo(message.chat.id, photo)
    photo.close()


message_counter = 0
@dp.message_handler()
async def mess_handler(message: types.Message):
    text0 = message.text.lower()

    global message_counter
    message_counter +=1


    try:
        try:
            user = message.from_user.username
        except:
            user = message.from_user.first_name
    except:
        user = 'Unnamed'
    userid = message.from_user.id
    text = text0.split()
    text = [re.sub(r'[^\w\s]', '', i) for i in text]

    # if '111' in text:
    #     await bot.restrict_chat_member(chat_id=message.chat.id, user_id=363700041,
    #                                    can_send_messages=True,
    #                                    can_send_other_messages=True,
    #                                    can_send_media_messages=True, can_add_web_page_previews=True)

    x= dict(message)
    jsoon = json.dumps(x, indent=4, ensure_ascii=False)


    for word in fuck.fuck_list:
        for word2 in fuck.fuck2_list:
            for mes in text:
                if (word == mes and 'spoiler' not in str(message.entities) or (word2 in mes and 'spoiler' not in str(message.entities))):

                    indx = text.index(mes)
                    lenofwrd = len(text[indx])
                    stars = lenofwrd - 2
                    text[indx] = text[indx][0] + ('*' * stars) + text[indx][-1]
                    censure_text = ' '.join(text)

                    bad_words = []
                    for i in fuck.fuck_list:
                        for j in fuck.fuck2_list:
                            for k in text:
                                if k == i or k == j:
                                    bad_words.append(k)
                    print(bad_words)

                    if len(bad_words) == 0:
                        await message.forward(chat_id=363700041)

                        await message.delete()

                        await bot.send_message(message.chat.id, f'@{user} сказал(а) так: {censure_text}')


    text0 = message.text.lower()
    text = normalize(text0)

    if 'id' in text0:
        await bot.send_message(message.chat.id,
                               f'Твой ID: {message.from_user.id}')


    if message_counter % 80 == 0:
        parser.whole_memes()
        photo = open('my_image.jpg', 'rb')
        await bot.send_photo(message.chat.id, photo)
        photo.close()



    # for hi in fuck.hello:
    #     if hi in text:
    #         await message.reply(random.choice(parser.bot_hi))

async def clearexcel():
    file = 'toban.xlsx'
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    wb.save(file)
    print('excel was cleared successfully')


async def scheduler():
    aioschedule.every(12).hours.do(clearexcel)
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(1)

async def on_startup(_):
    asyncio.create_task(scheduler())

if __name__== '__main__':
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
